
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
توابع کمکی برای ربات Excel
"""

import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from telegram import ReplyKeyboardMarkup, KeyboardButton

from config import DEFAULT_FIELDS, FIELDS_FILE, EXCEL_FILE, THEMES


def load_fields():
    """بارگذاری فیلدها از فایل"""
    try:
        if os.path.exists(FIELDS_FILE):
            with open(FIELDS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f).get('fields', DEFAULT_FIELDS)
    except Exception:
        pass
    return DEFAULT_FIELDS


def save_fields(fields):
    """ذخیره فیلدها در فایل"""
    try:
        with open(FIELDS_FILE, 'w', encoding='utf-8') as f:
            json.dump({'fields': fields}, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def load_user_theme(user_id):
    """بارگذاری تم کاربر"""
    try:
        theme_file = f"theme_{user_id}.json"
        if os.path.exists(theme_file):
            with open(theme_file, 'r', encoding='utf-8') as f:
                return json.load(f).get('theme', 'blue')
    except Exception:
        pass
    return 'blue'


def save_user_theme(user_id, theme):
    """ذخیره تم کاربر"""
    try:
        theme_file = f"theme_{user_id}.json"
        with open(theme_file, 'w', encoding='utf-8') as f:
            json.dump({'theme': theme}, f, ensure_ascii=False)
        return True
    except Exception:
        return False


def fix_card_numbers(df):
    """تصحیح شماره کارت ها و کدهای عددی"""
    df_copy = df.copy()
    for col in df_copy.columns:
        if any(x in col.lower() for x in ['شماره', 'کد', 'کارت', 'تلفن']):
            df_copy[col] = df_copy[col].astype(str)
            df_copy[col] = df_copy[col].str.replace('.0', '', regex=False)
            df_copy[col] = df_copy[col].replace('nan', '')
    return df_copy


def create_excel(df, theme="blue"):
    """ایجاد فایل Excel با فرمت زیبا"""
    try:
        # تصحیح شماره ها
        df_fixed = fix_card_numbers(df)
        
        # ایجاد workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        
        # اضافه کردن داده ها
        for r in dataframe_to_rows(df_fixed, index=False, header=True):
            ws.append(r)
        
        # اگر هیچ داده ای نیست، فقط هدرها را اضافه کن
        if df_fixed.empty:
            return wb.save(EXCEL_FILE)
        
        # دریافت تم رنگی
        theme_colors = THEMES.get(theme, THEMES['blue'])
        
        # فرمت دهی سرستون
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color=theme_colors["header"], 
            end_color=theme_colors["header"], 
            fill_type="solid"
        )
        
        for col in range(1, len(df_fixed.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # فرمت دهی ردیف ها
        for row in range(2, len(df_fixed) + 2):
            color = theme_colors["row1"] if row % 2 == 0 else theme_colors["row2"]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            for col in range(1, len(df_fixed.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
                
                # فرمت ویژه برای شماره ها
                column_name = df_fixed.columns[col-1]
                if any(x in column_name.lower() for x in ['شماره', 'کد', 'کارت', 'تلفن']):
                    cell.number_format = '@'
                    if cell.value:
                        cell.value = str(cell.value)
        
        # تنظیم عرض ستون ها
        for col_idx, col in enumerate(ws.columns, 0):
            if col_idx < len(df_fixed.columns):
                column_name = df_fixed.columns[col_idx]
                if any(x in column_name.lower() for x in ['شماره', 'کد', 'کارت']):
                    ws.column_dimensions[col[0].column_letter].width = 22
                else:
                    max_length = max(len(str(cell.value)) for cell in col if cell.value)
                    ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
        
        # ثابت کردن سرستون
        ws.freeze_panes = 'A2'
        
        # ذخیره فایل
        wb.save(EXCEL_FILE)
        return True
        
    except Exception:
        return False


def get_keyboard():
    """کیبورد اصلی"""
    keyboard = [
        ["➕ اضافه کردن", "📋 نمایش همه", "📁 دریافت فایل"],
        ["✏️ ویرایش", "🗑️ حذف", "🔍 جستجو"],
        ["⚙️ مدیریت فیلدها", "🎨 تغییر تم", "📊 آمار"],
        ["🧹 حذف همه", "ℹ️ راهنما"]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def ensure_excel_file():
    """اطمینان از وجود فایل Excel"""
    if not os.path.exists(EXCEL_FILE) or os.path.getsize(EXCEL_FILE) == 0:
        fields = load_fields()
        df = pd.DataFrame(columns=fields)
        create_excel(df, "blue")


def validate_field_input(field_name, value):
    """اعتبارسنجی ورودی فیلد"""
    value = value.strip()
    
    if not value:
        return False, "❌ مقدار نمی تواند خالی باشد"
    
    # بررسی فیلدهای عددی
    if any(x in field_name.lower() for x in ['شماره', 'کد', 'کارت']):
        if not value.isdigit():
            return False, "❌ لطفاً فقط عدد وارد کنید"
    
    # بررسی سن
    if 'سن' in field_name.lower():
        try:
            age = int(value)
            if age < 0 or age > 150:
                return False, "❌ سن باید بین 0 تا 150 باشد"
        except ValueError:
            return False, "❌ لطفاً یک عدد معتبر وارد کنید"
    
    return True, value


def format_record_display(df, max_records=10):
    """فرمت دهی نمایش رکوردها"""
    if df.empty:
        return "📭 هیچ رکوردی وجود ندارد."
    
    records_to_show = min(max_records, len(df))
    message = f"📋 **نمایش {records_to_show} رکورد از {len(df)} رکورد:**\n\n"
    
    for i in range(records_to_show):
        row = df.iloc[i]
        message += f"🔹 **رکورد {i+1}:**\n"
        for col in df.columns:
            if pd.notna(row[col]) and str(row[col]) != 'nan':
                value = str(row[col])
                if value.endswith('.0'):
                    value = value[:-2]
                message += f"  • {col}: {value}\n"
        message += "\n"
    
    if len(df) > max_records:
        message += f"... و {len(df) - max_records} رکورد دیگر"
    
    return message


def format_search_results(results, keyword, max_results=5):
    """فرمت دهی نتایج جستجو"""
    if results.empty:
        return f"🔍 **نتیجه جستجو:**\n❌ هیچ رکوردی با کلیدواژه '{keyword}' یافت نشد."
    
    msg = f"🔍 **نتیجه جستجو:**\n"
    msg += f"🎯 کلیدواژه: {keyword}\n"
    msg += f"📊 {len(results)} رکورد یافت شد:\n\n"
    
    for i, (idx, row) in enumerate(results.iterrows()):
        if i >= max_results:
            msg += f"... و {len(results) - max_results} نتیجه دیگر"
            break
        
        msg += f"🔹 **نتیجه {i+1}:**\n"
        for col in results.columns:
            if pd.notna(row[col]) and str(row[col]) != 'nan':
                value = str(row[col])
                if value.endswith('.0'):
                    value = value[:-2]
                # هایلایت کلیدواژه
                if keyword.lower() in value.lower():
                    value = value.replace(keyword, f"**{keyword}**")
                msg += f"  • {col}: {value}\n"
        msg += "\n"
    
    return msg


def get_file_size_string(file_path):
    """تبدیل سایز فایل به رشته قابل خواندن"""
    try:
        if not os.path.exists(file_path):
            return "0 بایت"
        
        size = os.path.getsize(file_path)
        if size < 1024:
            return f"{size} بایت"
        elif size < 1024 * 1024:
            return f"{size / 1024:.1f} کیلوبایت"
        else:
            return f"{size / (1024 * 1024):.1f} مگابایت"
    except Exception:
        return "نامشخص"


def clean_value(value):
    """پاکسازی مقدار برای نمایش"""
    if pd.isna(value):
        return ""
    
    value_str = str(value)
    if value_str.endswith('.0'):
        value_str = value_str[:-2]
    
    return value_str.replace('nan', '')

